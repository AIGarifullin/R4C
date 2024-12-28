import http
import json

from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook

from api.utils import end_date, start_date, validate_robot_data
from robots.models import Robot


@csrf_exempt   # Отключили механизм защиты от межсайтовой подделки запросов
def get_robots_list_or_create_robot(request):
    """Получение списка роботов и их создание."""
    if request.method == 'GET':
        robots = Robot.objects.all().values('id', 'serial', 'model',
                                            'version', 'created')
        robots_list = list(robots)
        return JsonResponse(
            robots_list, safe=False, status=http.HTTPStatus.OK
            )
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            is_valid, error_message = validate_robot_data(data)

            if not is_valid:
                return JsonResponse(
                    {'error': error_message},
                    status=http.HTTPStatus.BAD_REQUEST
                    )
            robot = Robot(
                serial=f"{data['model']}-{data['version']}",
                model=data['model'],
                version=data['version'],
                created=data['created'],
            )
            robot.save()
            return JsonResponse(
                {'message': 'Robot is created', 'robot_id': robot.id},
                status=http.HTTPStatus.CREATED
                )
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'},
                                status=http.HTTPStatus.BAD_REQUEST
                                )
    return JsonResponse(
        {'error': 'Only GET and POST requests are allowed.'},
        status=http.HTTPStatus.METHOD_NOT_ALLOWED
        )


def get_robot(request, robot_id):
    """Получение робота по ID."""
    try:
        robot = Robot.objects.get(id=robot_id)
        response = {
            'id': robot.id,
            'serial': robot.serial,
            'model': robot.model,
            'version': robot.version,
            'created': robot.created
        }
        return JsonResponse(response)
    except Robot.DoesNotExist:
        return JsonResponse(
            {'error': 'The robot was not found.'},
            status=http.HTTPStatus.NOT_FOUND
        )


def get_robot_report(request):
    """Создание Excel-файла со сводкой по суммарным показателям
    производства роботов за последнюю неделю."""
    robots = Robot.objects.filter(created__range=(start_date, end_date))
    file = Workbook()

    robot_sum = {}
    for robot in robots:
        model = robot.model
        version = robot.version
        if model not in robot_sum:
            robot_sum[model] = {}
        if version not in robot_sum[model]:
            robot_sum[model][version] = 0
        robot_sum[model][version] += 1

    for model, versions in robot_sum.items():
        sheet = file.create_sheet(title=f'Model {model}')
        sheet.append(['Модель', 'Версия', 'Количество за неделю'])

        for version, count in versions.items():
            sheet.append([model, version, count])
    if 'Sheet' in file.sheetnames:
        del file['Sheet']
    file_name = 'robot_report.xlsx'
    type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=type)
    response['Content-Disposition'] = f'attachment; filename={file_name}'

    file.save(response)
    return response
